"""Excel (.xlsx) import utilities for project activities."""

from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from openpyxl import Workbook, load_workbook

from .models import Activity

REQUIRED_COLUMN_COUNT = 3
HEADER_ALIASES = {
    'activity name',
    'total quantity',
    'unit',
    'activity',
    'quantity',
    'name',
}


def _build_unit_lookup():
    lookup = {}
    for code, label in Activity.UNIT_CHOICES:
        lookup[code.lower()] = code
        lookup[label.lower()] = code
        lookup[label.lower().replace(' ', '')] = code
    lookup.update({
        'square feet': 'sqft',
        'square foot': 'sqft',
        'sq ft': 'sqft',
        'square meters': 'sqm',
        'square metre': 'sqm',
        'square metres': 'sqm',
        'sq m': 'sqm',
        'cubic feet': 'cuft',
        'cubic meters': 'cum',
        'cubic metres': 'cum',
        'linear feet': 'lnft',
        'linear meters': 'lnm',
        'linear metres': 'lnm',
        'pieces': 'pcs',
        'piece': 'pcs',
        'kilogram': 'kg',
        'kilograms': 'kg',
        'tons': 'ton',
        'liter': 'ltr',
        'liters': 'ltr',
        'litre': 'ltr',
        'litres': 'ltr',
        'bag': 'bags',
        'hour': 'hours',
        'day': 'days',
    })
    return lookup


UNIT_LOOKUP = _build_unit_lookup()


def _cell_value(value):
    if value is None:
        return ''
    return str(value).strip()


def _is_header_row(cells):
    normalized = [_cell_value(c).lower() for c in cells[:3]]
    if not any(normalized):
        return False
    matches = sum(1 for value in normalized if value in HEADER_ALIASES)
    return matches >= 2


def _resolve_unit(raw_unit):
    key = _cell_value(raw_unit).lower()
    if not key:
        return None, 'Unit is required.'
    if key in UNIT_LOOKUP:
        return UNIT_LOOKUP[key], None
    compact = key.replace('.', '').replace(' ', '')
    if compact in UNIT_LOOKUP:
        return UNIT_LOOKUP[compact], None
    allowed = ', '.join(label for _, label in Activity.UNIT_CHOICES)
    return None, f'Invalid unit "{raw_unit}". Allowed values include: {allowed}.'


def _parse_quantity(raw_quantity):
    if raw_quantity is None or _cell_value(raw_quantity) == '':
        return None, 'Total Quantity is required.'

    try:
        quantity = Decimal(str(raw_quantity).strip().replace(',', ''))
    except (InvalidOperation, ValueError):
        return None, f'Total Quantity must be numeric (got "{raw_quantity}").'

    if quantity <= 0:
        return None, 'Total Quantity must be greater than zero.'

    return quantity, None


def validate_workbook_structure(workbook):
    if not workbook.sheetnames:
        return False, 'The Excel file contains no worksheets.'

    worksheet = workbook[workbook.sheetnames[0]]

    if worksheet.max_column != REQUIRED_COLUMN_COUNT:
        return (
            False,
            f'Invalid file structure: exactly {REQUIRED_COLUMN_COUNT} columns are required '
            f'(Activity Name, Total Quantity, Unit). Found {worksheet.max_column} column(s).',
        )

    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=3, values_only=True),
        start=1,
    ):
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        extra_values = [cell for cell in row[3:] if cell is not None and str(cell).strip() != '']
        if extra_values:
            return False, f'Row {row_index} contains data beyond column C. Only 3 columns are allowed.'

    return True, None


def build_sample_workbook():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Activities'
    worksheet.append(['Activity Name', 'Total Quantity', 'Unit'])
    worksheet.append(['Rc wall', 89, 'Square Meters'])
    worksheet.append(['Excavation for foundation', 78, 'Square Feet'])
    worksheet.append(['Steel Work', 120, 'sqft'])
    worksheet.column_dimensions['A'].width = 32
    worksheet.column_dimensions['B'].width = 18
    worksheet.column_dimensions['C'].width = 18

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def import_activities_from_xlsx(uploaded_file, project, created_by):
    """
    Parse and import activities from an Excel file.

    Returns a summary dict suitable for JSON responses.
    """
    summary = {
        'success': False,
        'file_error': None,
        'total_rows': 0,
        'successful': 0,
        'failed': 0,
        'failures': [],
        'created_count': 0,
        'message': '',
    }

    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception:
        summary['file_error'] = 'Unable to read the Excel file. The file may be corrupt or not a valid .xlsx file.'
        summary['message'] = summary['file_error']
        return summary

    is_valid, structure_error = validate_workbook_structure(workbook)
    if not is_valid:
        summary['file_error'] = structure_error
        summary['message'] = structure_error
        return summary

    worksheet = workbook[workbook.sheetnames[0]]
    parsed_rows = []
    skip_header = False

    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=3, values_only=True),
        start=1,
    ):
        name, quantity_raw, unit_raw = row
        if all(value is None or str(value).strip() == '' for value in row):
            continue

        if not skip_header and _is_header_row(row):
            skip_header = True
            continue

        parsed_rows.append({
            'row_number': row_index,
            'name': _cell_value(name),
            'quantity_raw': quantity_raw,
            'unit_raw': unit_raw,
        })

    summary['total_rows'] = len(parsed_rows)

    if not parsed_rows:
        summary['file_error'] = 'No activity rows found in the Excel file.'
        summary['message'] = summary['file_error']
        return summary

    existing_names = {
        name.lower()
        for name in Activity.objects.filter(project=project, is_active=True).values_list('name', flat=True)
    }
    seen_names = set()
    valid_records = []

    for row in parsed_rows:
        row_number = row['row_number']
        name = row['name']

        if not name:
            summary['failed'] += 1
            summary['failures'].append({'row': row_number, 'reason': 'Activity Name is required.'})
            continue

        quantity, quantity_error = _parse_quantity(row['quantity_raw'])
        if quantity_error:
            summary['failed'] += 1
            summary['failures'].append({'row': row_number, 'reason': quantity_error})
            continue

        unit_code, unit_error = _resolve_unit(row['unit_raw'])
        if unit_error:
            summary['failed'] += 1
            summary['failures'].append({'row': row_number, 'reason': unit_error})
            continue

        name_key = name.lower()
        if name_key in existing_names or name_key in seen_names:
            summary['failed'] += 1
            summary['failures'].append({
                'row': row_number,
                'reason': f'Activity "{name}" already exists for this project.',
            })
            continue

        seen_names.add(name_key)
        valid_records.append(
            Activity(
                project=project,
                name=name,
                total_quantity=quantity,
                unit=unit_code,
                created_by=created_by,
                is_active=True,
            )
        )

    if valid_records:
        try:
            with transaction.atomic():
                Activity.objects.bulk_create(valid_records)
            summary['successful'] = len(valid_records)
            summary['created_count'] = len(valid_records)
        except Exception:
            summary['failed'] += len(valid_records)
            summary['successful'] = 0
            summary['created_count'] = 0
            summary['failures'].append({
                'row': 'bulk',
                'reason': 'Database error while saving activities. No rows were imported.',
            })
            summary['message'] = 'Import failed during database save.'
            return summary

    summary['success'] = summary['successful'] > 0 or summary['total_rows'] == 0
    if summary['successful'] and summary['failed']:
        summary['message'] = (
            f'Import completed with partial success: {summary["successful"]} inserted, '
            f'{summary["failed"]} failed out of {summary["total_rows"]} rows.'
        )
    elif summary['successful']:
        summary['message'] = f'Successfully imported {summary["successful"]} activities.'
        summary['success'] = True
    else:
        summary['message'] = 'No activities were imported. Please review the failed rows.'
        summary['success'] = False

    return summary
